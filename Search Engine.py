# -*- coding: utf-8 -*-
"""Untitled1.ipynb

Automatically generated by Colaboratory.

Original file is located at
    https://colab.research.google.com/drive/1ZQzPgN1VDzts3f0K0mg3QCewhCd4M2wB
"""

!pip3 install newspaper3k

import nltk
from newspaper import Article
import openpyxl
print("Enter Keyword")
word = input().lower()

# import the URLs from the Excel
from openpyxl import load_workbook
wb = load_workbook(r'urls.xlsm') 
ws = wb.get_sheet_by_name('Sheet1')  
column = ws['A']  # Column
column_list = [column[x].value for x in range(len(column))] # create a list
url_list = list(filter(None, column_list)) # remove blanks
url_list.pop(0) # remove title

# start loop
x = 0
while x < len(url_list):


   url = str(url_list[x]) # set url  
   article = Article(url)
   article.download()
   article.parse()
   article.authors
   #print(article.publish_date)
   #print(article.text)
   #print(article.authors)
   #print(article.top_image)  
   c = article.text.lower()
   count = c.count(word)
   if count > 0:
     print("Here's your search result")
     print(article.title)
     print("Author Name",article.authors)
     print("Publication date",article.publish_date)
     print("Link",url)
     


   


   x = x + 1 # move to next url